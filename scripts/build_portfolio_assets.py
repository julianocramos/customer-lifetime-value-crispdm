from pathlib import Path
import textwrap
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
REPORT = ROOT / "reports" / "ltv_case_report.pdf"
EXCEL = ROOT / "deploy" / "ltv_operational_batch.xlsx"
REPORT.parent.mkdir(exist_ok=True)
EXCEL.parent.mkdir(exist_ok=True)

NAVY = "#17365D"
BLUE = "#4F81BD"
LIGHT = "#EAF2F8"
TEAL = "#2A9D8F"
ORANGE = "#F4A261"
RED = "#D9534F"
GRAY = "#666666"

TOTAL_CUSTOMERS = 38753
TOTAL_LTV = 84964816.22

CV = [
    ["Dummy", -0.0001, -0.0026, -0.0011, -0.0000, -0.0004, -0.0009],
    ["Linear", 0.8489, 0.8440, 0.8533, 0.8486, 0.8572, 0.8504],
    ["Poly (d=2)", 0.8434, 0.8399, 0.8501, 0.8455, 0.8533, 0.8464],
    ["RF", 0.8210, 0.8152, 0.8313, 0.8254, 0.8339, 0.8254],
]
TEST = [
    ["Dummy", -0.0001, 1301.31, 1024.58],
    ["Linear", 0.8478, 507.74, 404.03],
    ["Poly (d=2)", 0.8453, 511.88, 407.41],
    ["RF", 0.8229, 547.69, 435.37],
]
PRODUCTS = [
    ["Python",11444,2097.22,24000592.46],
    ["Community",10827,1778.27,19253351.38],
    ["Power BI",6971,2378.33,16579320.09],
    ["Excel",4767,2030.97,9681629.90],
    ["Lifetime Access",1610,4980.48,8018576.27],
    ["Commercial",862,2531.73,2182350.29],
    ["Other",870,2399.47,2087536.22],
    ["Data Science",419,2735.36,1146116.39],
    ["Full Stack JavaScript",501,2022.65,1013349.42],
    ["Data Analysis",482,2078.83,1001993.80],
]
CHANNELS = [
    ["Launch",29998,2188.38,65647040.96],
    ["Waitlist",3833,2365.55,9067165.71],
    ["Direct Traffic",1176,2347.56,2760735.62],
    ["Other",904,2523.59,2281326.77],
    ["Commercial",873,2516.30,2196726.84],
    ["Webinar",752,2168.37,1630617.80],
    ["Checkout",1217,1134.92,1381202.52],
]
RECURRENCE = [
    ["Non-recurring",22529,3004.93,67698074.34,1153.99],
    ["Recurring",16224,1064.27,17266741.88,108.20],
]

PARAMS = {
    "intercept": 2747.8860654835735,
    "mean_first_purchase": 715.1107747634687,
    "scale_first_purchase": 665.9064613725793,
    "coef_first_purchase": 1013.2873103571518,
    "coef_recurring": -368.7296916216356,
}
COEFS = {
    "entry_product": {
        "Commercial":0.0,"Community":-306.13934499170807,"Data Analysis":-155.9120117547081,
        "Data Science":91.26021762208455,"Excel":-253.36933078605395,
        "Full Stack JavaScript":-107.6436041491745,"Lifetime Access":354.02343969186836,
        "Other":-193.3406707514381,"Power BI":36.018635174877765,"Python":-157.04202564378025,
    },
    "sales_channel": {
        "Checkout":0.0,"Commercial":-138.72997200847198,"Direct Traffic":-528.614588328954,
        "Launch":-211.39336687705062,"Other":-614.4626460042267,
        "Waitlist":-412.80729030685717,"Webinar":-169.99012372592279,
    },
    "gender": {"Female":0.0,"Male":9.714148867380832,"Other":7.747227096851855},
    "education": {
        "Basic Education":0.0,"High School":-26.868855007722686,
        "Higher Education - Complete":-17.382228134447868,
        "Higher Education - Incomplete":-24.013084058009756,
        "Higher Education+":-23.84950788554162,"Not Reported":-27.92384865699283,
    },
    "purchase_month": {
        "1":0.0,"2":6.823636153664222,"3":7.9287154060584735,"4":11.615456819275723,
        "5":-25.72489275593819,"6":11.264137525624022,"7":-12.462186386139905,
        "8":-28.91453218494131,"9":23.72690737288671,"10":12.534944008951202,
        "11":22.69886648362883,"12":13.72426878021508,
    },
    "purchase_day_of_week": {
        "Friday":0.0,"Monday":-4.071061642523554,"Saturday":6.2257800857395,
        "Sunday":2.0766874335308954,"Thursday":-12.426676794120015,
        "Tuesday":9.222414228021588,"Wednesday":-6.058396875396041,
    },
}


def money(x):
    return f"R$ {x:,.2f}"


def add_banner(fig, title, subtitle=None):
    fig.text(0.055, 0.955, title, fontsize=22, fontweight="bold", color=NAVY, va="top")
    fig.patches.append(plt.Rectangle((0.055,0.905),0.89,0.008,transform=fig.transFigure,color=BLUE,ec=None))
    if subtitle:
        fig.text(0.055,0.885,subtitle,fontsize=10.5,color=GRAY,va="top")


def add_footer(fig, page):
    fig.text(0.055,0.035,"Customer Lifetime Value Prediction | CRISP-DM",fontsize=8,color=GRAY)
    fig.text(0.945,0.035,str(page),fontsize=8,color=GRAY,ha="right")


def draw_table(ax, headers, rows, bbox, colwidths, fontsize=8.5):
    table=ax.table(cellText=rows,colLabels=headers,cellLoc="center",loc="center",bbox=bbox,colWidths=colwidths)
    table.auto_set_font_size(False); table.set_fontsize(fontsize)
    for (r,c),cell in table.get_celld().items():
        cell.set_linewidth(0.6)
        if r==0:
            cell.set_facecolor(BLUE); cell.get_text().set_color("white"); cell.get_text().set_weight("bold")
        else:
            cell.set_facecolor("#F7FBFF" if r%2 else LIGHT)
            if c==0: cell.get_text().set_weight("bold"); cell.get_text().set_color(NAVY)
    return table


def smart_scatter(ax, data, title, xlabel="Number of customers", ylabel="Average LTV (R$)"):
    names=[r[0] for r in data]; x=np.array([r[1] for r in data],float); y=np.array([r[2] for r in data],float)
    ax.scatter(x,y,s=75,color=TEAL,zorder=3)
    ax.set_title(title,fontsize=14,fontweight="bold",color=NAVY,pad=12)
    ax.set_xlabel(xlabel); ax.set_ylabel(ylabel); ax.grid(alpha=.2,zorder=0)
    offsets=[(10,10),(10,-16),(-70,10),(-60,-18),(10,10),(10,12),(10,-18),(-75,12),(-62,-18),(10,12)]
    for i,(name,xi,yi) in enumerate(zip(names,x,y)):
        dx,dy=offsets[i%len(offsets)]
        ax.annotate(name,(xi,yi),xytext=(dx,dy),textcoords="offset points",fontsize=8.5,
                    bbox=dict(boxstyle="round,pad=.18",fc="white",ec="#CCCCCC",alpha=.95),
                    arrowprops=dict(arrowstyle="-",color="#AAAAAA",lw=.6))


def build_pdf():
    with PdfPages(REPORT) as pdf:
        # 1 Cover
        fig=plt.figure(figsize=(8.27,11.69)); fig.patch.set_facecolor("white"); plt.axis("off")
        fig.patches.append(plt.Rectangle((0,0.91),1,.09,transform=fig.transFigure,color=NAVY,ec=None))
        fig.text(.055,.955,"Customer Lifetime Value Prediction",fontsize=24,fontweight="bold",color="white",va="center")
        fig.text(.055,.895,"CRISP-DM Portfolio Report",fontsize=12,color=BLUE,va="top")
        cards=[("Customers",f"{TOTAL_CUSTOMERS:,}"),("Selected Model","Linear"),("Test R²","0.8478")]
        for i,(k,v) in enumerate(cards):
            x=.055+i*.30
            fig.patches.append(plt.Rectangle((x,.73),.27,.11,transform=fig.transFigure,facecolor=LIGHT,edgecolor=BLUE,lw=1.2))
            fig.text(x+.015,.805,k,fontsize=9,color=BLUE,fontweight="bold")
            fig.text(x+.015,.765,v,fontsize=18,color=NAVY,fontweight="bold")
        fig.text(.055,.64,"Business objective",fontsize=16,fontweight="bold",color=NAVY)
        objective=("Estimate Customer Lifetime Value using only information available at the first purchase, "
                   "supporting acquisition-budget allocation, product and channel prioritization, segmentation, and CAC decisions.")
        fig.text(.055,.60,textwrap.fill(objective,88),fontsize=11,color="#222",va="top")
        fig.text(.055,.48,"Test-set performance",fontsize=16,fontweight="bold",color=NAVY)
        for j,s in enumerate(["R²  0.8478","RMSE  R$ 507.74","MAE  R$ 404.03","41 transformed features"]):
            fig.text(.075,.435-j*.045,"• "+s,fontsize=11,color="#222")
        fig.text(.055,.19,"38,753 customers  |  Historical LTV represented: R$ 84,964,816.22",fontsize=10,color=GRAY)
        fig.text(.055,.145,"All figures in this report come from the executed Python workflow and the prepared modeling dataset.",fontsize=9,color=GRAY)
        add_footer(fig,1); pdf.savefig(fig,bbox_inches="tight"); plt.close(fig)

        # 2 Models
        fig,ax=plt.subplots(figsize=(11.69,8.27)); ax.axis("off"); add_banner(fig,"Model Evaluation","Five-fold cross-validation on the training set and held-out test performance")
        cvrows=[[r[0]]+[f"{v:.4f}" for v in r[1:]] for r in CV]
        draw_table(ax,["Model","Fold 1","Fold 2","Fold 3","Fold 4","Fold 5","Mean R²"],cvrows,[.055,.54,.89,.27],[.19,.12,.12,.12,.12,.12,.13],9)
        testrows=[[r[0],f"{r[1]:.4f}",money(r[2]),money(r[3])] for r in TEST]
        draw_table(ax,["Model","Test R²","RMSE","MAE"],testrows,[.13,.25,.74,.20],[.28,.18,.27,.27],9)
        fig.text(.065,.15,"Linear Regression was selected: strongest validation performance with high interpretability.",fontsize=11,color=NAVY,fontweight="bold")
        add_footer(fig,2); pdf.savefig(fig,bbox_inches="tight"); plt.close(fig)

        # 3 Products
        fig=plt.figure(figsize=(11.69,8.27)); add_banner(fig,"Entry Product Analysis","Average customer value must be interpreted together with segment scale")
        ax1=fig.add_axes([.07,.50,.40,.35]); ax2=fig.add_axes([.55,.50,.39,.35]); ax3=fig.add_axes([.06,.10,.88,.30]); ax3.axis("off")
        smart_scatter(ax1,PRODUCTS,"Average LTV vs Customer Volume")
        top=sorted(PRODUCTS,key=lambda r:r[3],reverse=True)
        names=[r[0] for r in top]; vals=[r[3]/1e6 for r in top]
        y=np.arange(len(names)); ax2.barh(y,vals,color=BLUE); ax2.set_yticks(y,names,fontsize=8); ax2.invert_yaxis(); ax2.set_xlabel("Total historical LTV (R$ millions)"); ax2.set_title("Total Historical LTV",fontsize=14,fontweight="bold",color=NAVY); ax2.grid(axis="x",alpha=.2)
        for yi,v in zip(y,vals): ax2.text(v+.15,yi,f"R$ {v:.2f}M",va="center",fontsize=7.8)
        rows=[]
        for n,c,a,t in PRODUCTS:
            rows.append([n,f"{c:,}",f"{c/TOTAL_CUSTOMERS:.2%}",money(a),money(t),f"{t/TOTAL_LTV:.2%}"])
        draw_table(ax3,["Entry Product","Customers","Share","Average LTV","Total LTV","LTV Share"],rows,[0,0,1,1],[.23,.11,.11,.16,.21,.13],7.6)
        fig.text(.07,.055,"Key insight: Lifetime Access leads in LTV per customer; Python leads in total historical LTV because of scale.",fontsize=9.5,color=NAVY,fontweight="bold")
        add_footer(fig,3); pdf.savefig(fig,bbox_inches="tight"); plt.close(fig)

        # 4 Channels
        fig=plt.figure(figsize=(11.69,8.27)); add_banner(fig,"Sales Channel Analysis","Scale changes the interpretation of average LTV")
        ax1=fig.add_axes([.07,.50,.40,.35]); ax2=fig.add_axes([.55,.50,.39,.35]); ax3=fig.add_axes([.07,.11,.87,.28]); ax3.axis("off")
        # manual channel offsets avoid Other/Commercial overlap
        names=[r[0] for r in CHANNELS]; x=np.array([r[1] for r in CHANNELS]); yv=np.array([r[2] for r in CHANNELS])
        ax1.scatter(x,yv,s=80,color=ORANGE,zorder=3); ax1.grid(alpha=.2); ax1.set_xlabel("Number of customers"); ax1.set_ylabel("Average LTV (R$)"); ax1.set_title("Average LTV vs Customer Volume",fontsize=14,fontweight="bold",color=NAVY)
        offsets={"Launch":(8,8),"Waitlist":(8,10),"Direct Traffic":(8,-16),"Other":(10,16),"Commercial":(10,-20),"Webinar":(8,8),"Checkout":(8,-16)}
        for n,xi,yi in zip(names,x,yv):
            dx,dy=offsets[n]; ax1.annotate(n,(xi,yi),xytext=(dx,dy),textcoords="offset points",fontsize=8.5,bbox=dict(boxstyle="round,pad=.18",fc="white",ec="#CCCCCC",alpha=.95),arrowprops=dict(arrowstyle="-",color="#AAA",lw=.6))
        ordered=sorted(CHANNELS,key=lambda r:r[3],reverse=True); n2=[r[0] for r in ordered]; v2=[r[3]/1e6 for r in ordered]; yy=np.arange(len(n2)); ax2.barh(yy,v2,color=TEAL); ax2.set_yticks(yy,n2,fontsize=8); ax2.invert_yaxis(); ax2.set_xlabel("Total historical LTV (R$ millions)"); ax2.set_title("Total Historical LTV",fontsize=14,fontweight="bold",color=NAVY); ax2.grid(axis="x",alpha=.2)
        for yi,v in zip(yy,v2): ax2.text(v+.3,yi,f"R$ {v:.2f}M",va="center",fontsize=8)
        rows=[[n,f"{c:,}",f"{c/TOTAL_CUSTOMERS:.2%}",money(a),money(t),f"{t/TOTAL_LTV:.2%}"] for n,c,a,t in CHANNELS]
        draw_table(ax3,["Sales Channel","Customers","Share","Average LTV","Total LTV","LTV Share"],rows,[0,0,1,1],[.20,.11,.11,.17,.23,.13],8)
        fig.text(.07,.055,"Key insight: Launch dominates total value because it represents more than three quarters of customers.",fontsize=9.5,color=NAVY,fontweight="bold")
        add_footer(fig,4); pdf.savefig(fig,bbox_inches="tight"); plt.close(fig)

        # 5 Recurrence
        fig=plt.figure(figsize=(11.69,8.27)); add_banner(fig,"Recurring vs Non-Recurring First Purchase","Historical association, not a causal conclusion")
        ax=fig.add_axes([.10,.43,.80,.40]); x=np.arange(2); width=.34
        customers=np.array([r[1] for r in RECURRENCE]); avg=np.array([r[2] for r in RECURRENCE]); labels=[r[0] for r in RECURRENCE]
        b1=ax.bar(x-width/2,customers,width,color=BLUE,label="Customers"); ax2=ax.twinx(); b2=ax2.bar(x+width/2,avg,width,color=ORANGE,label="Average LTV")
        ax.set_xticks(x,labels); ax.set_ylabel("Customers"); ax2.set_ylabel("Average LTV (R$)"); ax.grid(axis="y",alpha=.2)
        for b,v in zip(b1,customers): ax.text(b.get_x()+b.get_width()/2,b.get_height()+300,f"{v:,}",ha="center",fontsize=9,fontweight="bold")
        for b,v in zip(b2,avg): ax2.text(b.get_x()+b.get_width()/2,b.get_height()+40,money(v),ha="center",fontsize=9,fontweight="bold")
        ax3=fig.add_axes([.12,.16,.76,.16]); ax3.axis("off")
        rows=[[r[0],f"{r[1]:,}",money(r[2]),money(r[3]),money(r[4])] for r in RECURRENCE]
        draw_table(ax3,["Group","Customers","Average LTV","Total LTV","Avg. First Purchase"],rows,[0,0,1,1],[.20,.14,.18,.25,.20],9)
        fig.text(.08,.08,"Non-recurring customers have much higher observed LTV and first-purchase value. Observation window and LTV definition matter.",fontsize=9.5,color=NAVY,fontweight="bold")
        add_footer(fig,5); pdf.savefig(fig,bbox_inches="tight"); plt.close(fig)

        # 6 Takeaways
        fig=plt.figure(figsize=(8.27,11.69)); plt.axis("off"); add_banner(fig,"Final Business Takeaways","How the model can support acquisition decisions")
        takeaways=[
            ("1","Linear Regression is the selected model","Test R² 0.8478 | RMSE R$ 507.74 | MAE R$ 404.03"),
            ("2","LTV per customer and segment volume tell different stories","Lifetime Access leads per customer; Python leads in total value."),
            ("3","Launch is the economically dominant sales channel","29,998 customers and R$ 65.65M in historical LTV."),
            ("4","Predicted LTV is decision support, not a guaranteed outcome","Combine predicted LTV with CAC, contribution margin, payback, risk and uncertainty."),
            ("5","Production improvement: temporal validation","The current project uses a random 80/20 split plus 5-fold CV; temporal validation is the next step."),
        ]
        y=.78
        for num,title,body in takeaways:
            fig.patches.append(plt.Rectangle((.07,y-.055),.075,.075,transform=fig.transFigure,facecolor=BLUE,edgecolor="none"))
            fig.text(.107,y-.017,num,fontsize=18,fontweight="bold",color="white",ha="center",va="center")
            fig.text(.17,y,title,fontsize=12,fontweight="bold",color=NAVY,va="top")
            fig.text(.17,y-.035,textwrap.fill(body,70),fontsize=10,color="#333",va="top")
            y-=.14
        fig.text(.07,.10,"Deployment artifacts: explained Excel simulator + operational batch scoring workbook.",fontsize=10,color=TEAL,fontweight="bold")
        add_footer(fig,6); pdf.savefig(fig,bbox_inches="tight"); plt.close(fig)


def build_excel():
    wb=Workbook(); ws=wb.active; ws.title="New Customers"; params=wb.create_sheet("Model Parameters")
    thin=Side(style="thin",color="D9E2F3")
    header_fill=PatternFill("solid",fgColor="17365D"); sub_fill=PatternFill("solid",fgColor="4F81BD"); input_fill=PatternFill("solid",fgColor="EAF2F8")
    green_fill=PatternFill("solid",fgColor="C6EFCE"); red_fill=PatternFill("solid",fgColor="FFC7CE")
    ws.merge_cells("A1:S2"); ws["A1"]="OPERATIONAL LTV SCORING TABLE"; ws["A1"].font=Font(bold=True,color="FFFFFF",size=18); ws["A1"].fill=header_fill; ws["A1"].alignment=Alignment(horizontal="center",vertical="center")
    ws.merge_cells("A3:S3"); ws["A3"]="Enter first-purchase information in columns A:I. Predicted LTV and feature impacts calculate automatically."; ws["A3"].alignment=Alignment(horizontal="center"); ws["A3"].fill=PatternFill("solid",fgColor="D9EAF7")
    for c,v in zip("ABCD",["Model","Test R²","RMSE","MAE"]): ws[f"{c}5"]=v; ws[f"{c}5"].fill=sub_fill; ws[f"{c}5"].font=Font(bold=True,color="FFFFFF")
    ws.append([])
    ws["A6"]="Linear Regression"; ws["B6"]=0.8477576423540136; ws["C6"]=507.7353021589962; ws["D6"]=404.0261679823802
    headers=["Customer ID","First Purchase Value (R$)","Recurring?","Entry Product","Sales Channel","Gender","Education","Purchase Month","Purchase Day","Predicted LTV (R$)","First Purchase Impact","Recurrence Impact","Product Impact","Channel Impact","Gender Impact","Education Impact","Month Impact","Day Impact","Status"]
    for i,h in enumerate(headers,1):
        cell=ws.cell(8,i,h); cell.fill=header_fill; cell.font=Font(bold=True,color="FFFFFF"); cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    example=["EXAMPLE-001",997,"Yes","Commercial","Waitlist","Male","Higher Education - Complete",1,"Tuesday"]
    for i,v in enumerate(example,1): ws.cell(9,i,v)
    for r in range(10,109): ws.cell(r,1,f"CUST-{r-8:03d}")
    for row in ws.iter_rows(min_row=9,max_row=108,min_col=1,max_col=9):
        for cell in row: cell.fill=input_fill
    # Parameter sheet
    params.append(["Feature","Category","Coefficient","Note","","Parameter","Value"])
    for cell in params[1]: cell.fill=sub_fill; cell.font=Font(bold=True,color="FFFFFF")
    for feature,mapping in COEFS.items():
        first=True
        for cat,coef in mapping.items():
            params.append([feature,cat,coef,"Reference category" if first and coef==0 else "",None,None,None]); first=False
    params[2][5]="Intercept"; params[2][6]=PARAMS["intercept"]
    params[3][5]="Mean - first_purchase_value"; params[3][6]=PARAMS["mean_first_purchase"]
    params[4][5]="Scale - first_purchase_value"; params[4][6]=PARAMS["scale_first_purchase"]
    params[5][5]="Coefficient - first_purchase_value"; params[5][6]=PARAMS["coef_first_purchase"]
    params[6][5]="Coefficient - recurring_first_purchase"; params[6][6]=PARAMS["coef_recurring"]
    maxp=params.max_row
    # Data validation
    def listdv(cells,values):
        dv=DataValidation(type="list",formula1='"'+','.join(str(v) for v in values)+'"'); ws.add_data_validation(dv); dv.add(cells)
    listdv("C9:C108",["No","Yes"]); listdv("D9:D108",list(COEFS["entry_product"])); listdv("E9:E108",list(COEFS["sales_channel"])); listdv("F9:F108",list(COEFS["gender"])); listdv("G9:G108",list(COEFS["education"])); listdv("H9:H108",range(1,13)); listdv("I9:I108",list(COEFS["purchase_day_of_week"]))
    # Formulas
    for r in range(9,109):
        lookup=lambda feature,cell: f'SUMIFS(\'Model Parameters\'!$C$2:$C${maxp},\'Model Parameters\'!$A$2:$A${maxp},"{feature}",\'Model Parameters\'!$B$2:$B${maxp},{cell})'
        fp=f"'Model Parameters'!$G$5*((B{r}-'Model Parameters'!$G$3)/'Model Parameters'!$G$4)"
        rec=f"'Model Parameters'!$G$6*IF(C{r}=\"Yes\",1,0)"
        prod=lookup("entry_product",f"D{r}"); chan=lookup("sales_channel",f"E{r}"); gender=lookup("gender",f"F{r}"); edu=lookup("education",f"G{r}")
        month=f'SUMIFS(\'Model Parameters\'!$C$2:$C${maxp},\'Model Parameters\'!$A$2:$A${maxp},"purchase_month",\'Model Parameters\'!$B$2:$B${maxp},TEXT(H{r},"0"))'
        day=lookup("purchase_day_of_week",f"I{r}")
        ws[f"J{r}"]=f'=IF(COUNTA(B{r}:I{r})<8,"",ROUND(\'Model Parameters\'!$G$2+{fp}+{rec}+{prod}+{chan}+{gender}+{edu}+{month}+{day},2))'
        for col,formula in zip("KLMNOPQR",[fp,rec,prod,chan,gender,edu,month,day]): ws[f"{col}{r}"]=f'=IF(B{r}="","",{formula})'
        ws[f"S{r}"]=f'=IF(COUNTA(B{r}:I{r})=0,"Waiting for input",IF(COUNTA(B{r}:I{r})<8,"Incomplete","Calculated"))'
    for c in range(10,19):
        for r in range(9,109): ws.cell(r,c).number_format='R$ #,##0.00;[Red]-R$ #,##0.00'
    ws["B6"].number_format="0.0000"; ws["C6"].number_format=ws["D6"].number_format='R$ #,##0.00'
    ws["B9"].number_format='R$ #,##0.00'
    ws.conditional_formatting.add("K9:R108",CellIsRule(operator="greaterThan",formula=["0"],fill=green_fill))
    ws.conditional_formatting.add("K9:R108",CellIsRule(operator="lessThan",formula=["0"],fill=red_fill))
    ws.conditional_formatting.add("S9:S108",FormulaRule(formula=['S9="Calculated"'],fill=green_fill))
    widths=[15,21,14,22,18,14,30,15,18,19,20,18,18,18,16,18,16,16,18]
    for i,w in enumerate(widths,1): ws.column_dimensions[chr(64+i)].width=w
    ws.freeze_panes="A9"; ws.auto_filter.ref="A8:S108"
    tab=Table(displayName="NewCustomerScoring",ref="A8:S108"); tab.tableStyleInfo=TableStyleInfo(name="TableStyleMedium2",showRowStripes=True,showColumnStripes=False); ws.add_table(tab)
    for col in "ABCD": ws[f"{col}5"].alignment=Alignment(horizontal="center")
    for col in range(1,8): params.column_dimensions[chr(64+col)].width=[28,32,18,22,4,32,20][col-1]
    params.freeze_panes="A2"
    wb.save(EXCEL)


def update_readme():
    p=ROOT/"README.md"
    text=p.read_text(encoding="utf-8")
    text=text.replace("images/model_comparison.png","images/model_comparison.svg")
    text=text.replace("images/ltv_simulator_preview.png","images/ltv_simulator_preview.svg")
    text=text.replace("images/ltv_vs_volume_product.png","images/ltv_vs_volume_product.svg")
    p.write_text(text,encoding="utf-8")


if __name__ == "__main__":
    build_pdf(); build_excel(); update_readme()
    print(REPORT, REPORT.stat().st_size)
    print(EXCEL, EXCEL.stat().st_size)
